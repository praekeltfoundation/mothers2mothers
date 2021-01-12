
from openpyxl import load_workbook
import json
import requests
import os
FILENAME = "who_content.xlsx"
SHEETNAME = "ImportInfo"
ZAMBIA_TOKEN = os.environ.get("ZAMBIA_TEAM_TOKEN")


def strip_language(keyword):
    return keyword.split("_", 1)[-1]

def get_media():
    response = requests.get(
        "https://whatsapp.turn.io/v1/export",
        headers={
            "Authorization": f"Bearer {ZAMBIA_TOKEN}",
            "Accept": "application/vnd.v1+json"
        }
    )
    response.raise_for_status()
    media = {}
    for content in response.json()["data"]:
        if content["attachment_media_object"]:
            media[strip_language(content["question"])] = content
    return media

media = get_media()

wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]
sheets = {}
all_languages = {}
uniq_languages = []
for row in ws.iter_rows(max_col=5):
    if row[1].value and str(row[0].value).lower() != "sheet":
        name = str(row[0].value).replace(" ", "_").replace("(", "").replace(")", "")
        sheets[f"{row[2].value}_{name}_{row[1].value}"] = {
            "sheet_name": row[0].value,
            "number_type": row[1].value,
            "country_name": row[2].value,
            "replace": {
                "South Africa": str(row[2].value),
                "SOUTH AFRICA": str(row[2].value).upper(),
                "Mozambique": str(row[2].value),
                "MOZAMBIQUE": str(row[2].value).upper(),
                "27600109000": str(row[3].value).replace("=", ""),
            },
        }
        language = str(row[0].value).split(" (")[0]
        if row[2].value not in all_languages:
            all_languages[row[2].value] = []
        if language not in all_languages[row[2].value]:
            all_languages[row[2].value].append(language)
        if language not in uniq_languages:
            uniq_languages.append(language)
# print(json.dumps(all_languages, indent=4))
def clean_keyword(word):
    word = word.strip()
    try:
        word = int(float(word))
    except:
        pass
    return str(word)
def process_automators(keywords, language):
    triggers = []
    if keywords:
        keywords = [clean_keyword(keyword) for keyword in str(keywords).split(",")]
        language_match = None
        if language != "no-lang":
            language_match = language.upper()
        if len(keywords) == 1 and keywords[0] == "CATCHALL":
            triggers.append(
                {
                    "trigger_params": {
                        "contact_field_match": "exact",
                        "contact_field_name": "language",
                        "contact_field_value": language_match,
                    },
                    "trigger_type": "message_inbound_catch_all",
                }
            )
        else:
            triggers.append(
                {
                    "trigger_params": {
                        "contact_field_match": "exact",
                        "contact_field_name": "language",
                        "contact_field_value": language_match,
                    },
                    "trigger_type": "message_inbound",
                }
            )
            if len(keywords) == 1:
                triggers.append(
                    {
                        "trigger_params": {"exact_match": keywords[0]},
                        "trigger_type": "message_inbound",
                    }
                )
            else:
                triggers.append(
                    {
                        "trigger_params": {"exact_matches": keywords},
                        "trigger_type": "message_inbound",
                    }
                )
        return triggers
    else:
        return []
def process_language_automations(keywords, language):
    automation = None
    if keywords:
        keywords = [clean_keyword(keyword) for keyword in str(keywords).split(",")]
        if len(keywords) == 1:
            params = {"exact_match": keywords[0]}
        else:
            params = {"exact_matches": keywords}
        automation = {
            "config": {
                "data": {
                    "actions": [
                        {"action_type": "reply"},
                        {
                            "action_params": {
                                "contact_field_name": "language",
                                "contact_field_value": language.upper(),
                            },
                            "action_type": "update_contact",
                        },
                    ],
                    "operator": "AND",
                    "triggers": [
                        {
                            "trigger_params": params,
                            "trigger_type": "message_inbound",
                        }
                    ],
                },
                "version": "0.2.0",
            },
            "enabled": True,
            "is_deleted": False,
            "name": f"{language}-language-switch",
        }
    return automation
def replace_content_values(content, values):
    for from_value, to_value in values["replace"].items():
        content = content.replace(from_value, to_value)
    return content
def replace_content_languages(content_name, content, languages):
    if content_name == "eng_language":
        parts = content.split("\n\n")
        languages = "\n".join(languages)
        return "\n\n".join([parts[0], parts[1], languages, parts[3]])
    return content
content_count = 0
automation_count = 0
number_data = {}
for key, values in sheets.items():
    sheet_name = values["sheet_name"]
    country_name = values["country_name"]
    number_type = values["number_type"]
    number_desc = f"content_{country_name}_{number_type}.json"
    if number_desc not in number_data:
        number_data[number_desc] = []
    print(f"processing: {key}")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_col=6):
        if row[1].value:
            question = row[0].value
            content = row[1].value
            language = row[2].value
            automation_triggers = process_automators(row[3].value, language)
            language_automation = process_language_automations(row[4].value, language)
            automators = []
            if automation_triggers:
                automators.append(
                    {
                        "config": {
                            "data": {
                                "actions": [{"action_type": "reply"}],
                                "operator": "AND",
                                "triggers": automation_triggers,
                            },
                            "version": "0.2.0",
                        },
                        "enabled": True,
                        "is_deleted": False,
                        "name": question.replace("_", "-"),
                    }
                )
                if language == "eng":
                    automation_triggers = process_automators(row[3].value, "no-lang")
                    if automation_triggers:
                        automators.append(
                            {
                                "config": {
                                    "data": {
                                        "actions": [{"action_type": "reply"}],
                                        "operator": "AND",
                                        "triggers": automation_triggers,
                                    },
                                    "version": "0.2.0",
                                },
                                "enabled": True,
                                "is_deleted": False,
                                "name": question.replace("_", "-").replace("eng", "null-language"),
                            }
                        )
            if language_automation:
                automators.append(language_automation)
            content = replace_content_values(content, values)
            content = replace_content_languages(question, content, all_languages[country_name])
            media_content = media.get(strip_language(question), {})
            number_data[number_desc].append(
                {
                    "answer": content,
                    "attachment_media_object": media_content.get("attachment_media_object"),
                    "attachment_media_type": media_content.get("attachment_media_type"),
                    "attachment_mime_type": media_content.get("attachment_mime_type"),
                    "attachment_uri": media_content.get("attachment_uri"),
                    "automators": automators,
                    "is_deleted": False,
                    "language": language,
                    "question": question,
                }
            )
            content_count += 1
            automation_count += len(automators)
for output_file, data in number_data.items():
    output = {"data": data, "schema_version": "0.1"}
    with open(output_file.replace(" ", "_"), "w") as outfile:
        json.dump(output, outfile, indent=4)
print("done")
print("")
print(f"Turn Numbers:  {len(number_data.keys())}")
print(f"Languages:     {len(uniq_languages)}")
print(f"Sheets:        {len(sheets.keys())}")
print(f"Content:       {content_count}")
print(f"Automators:    {automation_count}")
