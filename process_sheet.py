import re
from functools import lru_cache

import emoji
from openpyxl import load_workbook

FILENAME = "who_content.xlsx"

global is_error
is_error = False


def error(message):
    """
    Prints out the error, and marks that we have an error. This allows us to print out
    all errors, before raising an exception
    """
    global is_error
    is_error = True
    print(message)


def base_emoji(keyword):
    """
    Strips any modifiers from the emoji, and just returns the base emoji
    """
    modifiers = [
        # We only remove skin type modifiers here, variation selectors and gender
        # modifiers aren't recognised the same as skin type modifiers by Turn, so
        # those will have to be specified as separate keywords.
        "\U0001f3fb",  # skin type 1-2
        "\U0001f3fc",  # skin type 3
        "\U0001f3fd",  # skin type 4
        "\U0001f3fe",  # skin type 5
        "\U0001f3ff",  # skin type 6
    ]
    if emoji.get_emoji_regexp().match(keyword):
        for m in modifiers:
            keyword = keyword.replace(m, "")

    return keyword


@lru_cache(maxsize=None)
def get_index(sheet, *names):
    """
    Returns the column index for the first matching name, assuming first row is header
    Matching is done with leading and trailing whitespace stripped, and case insensitive
    """
    header = [(c.value or "").strip().lower() for c in sheet[1]]
    for name in names:
        name = name.strip().lower()
        if name in header:
            return header.index(name)
    raise AttributeError(f"None of {names} found in header {header}")


def get_cell(sheet, row, name):
    """
    Gets the value from the column `name` from the `row` in `sheet`
    """
    column_names = {
        "content title": ["content_title", "content title", "automation title"],
        "content": ["content"],
        "language": ["language"],
        "automation": ["automation"],
    }[name]
    return row[get_index(sheet, *column_names)]


def get_keywords(sheet):
    """
    Returns a dictionary of content name and a list of keywords
    """

    def get_content_title(row):
        content = get_cell(sheet, row, "content title").value or ""
        language = get_cell(sheet, row, "language").value
        if content.startswith(language):
            return content[len(language) + 1 :]
        return content

    def get_keywords(row):
        automation = get_cell(sheet, row, "automation").value or ""
        return automation.split(",")

    return {
        get_content_title(row): get_keywords(row)
        for row in sheet.iter_rows(min_row=2)
        if get_cell(sheet, row, "content title").value
    }


def get_content(sheet):
    """
    Returns a dictionary of content name and content
    """

    def get_content_title(row):
        content = get_cell(sheet, row, "content title").value or ""
        language = get_cell(sheet, row, "language").value
        if content.startswith(language):
            return content[len(language) + 1 :]
        return content

    return {
        get_content_title(row): get_cell(sheet, row, "content").value
        for row in sheet.iter_rows(min_row=2)
        if get_cell(sheet, row, "content title").value
        and get_cell(sheet, row, "content").value
    }


def clean_keywords(workbook):
    """
    Goes through all of the content sheets, and cleans up the keywords, by:
    - Removing whitespace
    - Removing empty and duplicate
    - Removing emoji modifiers
    - Error on keyword that contains more than just an emoji
    - Error on duplicate across rows in same sheet
    """
    for sheet in workbook:
        if sheet.title.strip().lower() in ("language codes", "importinfo"):
            continue

        seen_keywords = set()
        for row in sheet.iter_rows(min_row=2):
            keywords = get_cell(sheet, row, "automation").value
            # Convert to string from different types
            if keywords is None:
                keywords = ""
            elif isinstance(keywords, float):
                keywords = str(int(keywords))
            elif isinstance(keywords, int):
                keywords = str(keywords)
            # Remove whitespace
            keywords = [k.strip() for k in keywords.split(",")]
            # Remove empty
            keywords = [k for k in keywords if k]
            # Remove emoji modifiers
            keywords = [base_emoji(k) for k in keywords]
            # Check for emoji keywords that have more than just the emoji in them
            for keyword in keywords:
                # Remove variation mods, since they're seen as separate by the regex
                keyword = keyword.replace("\ufe0f", "")
                keyword = keyword.replace("\ufe0e", "")
                match = emoji.get_emoji_regexp().match(keyword)
                if match and len(match[0]) < len(keyword):
                    error(
                        f"Invalid keyword, more than just emoji: {keyword}; "
                        f"sheet: {sheet.title}"
                    )
            # Remove duplicate in row
            deduped = []
            for keyword in keywords:
                if keyword not in deduped:
                    deduped.append(keyword)
            keywords = deduped
            # Check for duplicate keywords across sheet
            for keyword in keywords:
                # Skip myths, that's allowed to be duplicated
                if "myths" in get_cell(sheet, row, "content title").value:
                    continue
                if keyword in seen_keywords:
                    error(f"Duplicate keyword {keyword}; sheet: {sheet.title}")
                seen_keywords.add(keyword)
            get_cell(sheet, row, "automation").value = ",".join(keywords)


def clean_language(workbook):
    """
    Goes through all the content sheets, and ensures that there's a language field
    present. If not, fills in using the last value for the language
    """
    for sheet in workbook:
        if sheet.title.strip().lower() in (
            "language codes",
            "importinfo",
        ):
            continue

        lang = None
        for row in sheet.iter_rows(min_row=2):
            language = get_cell(sheet, row, "language")
            content_title = get_cell(sheet, row, "content title")
            if content_title.value:
                if language.value:
                    lang = language.value
                else:
                    language.value = lang


def clean_content_title(workbook):
    """
    Goes through all the content sheets, and normalises the content title:
    - Whitespace trimmed from start and finish
    - Non word characters replace with `-`
    """
    for sheet in workbook:
        if sheet.title.strip().lower() in (
            "language codes",
            "importinfo",
        ):
            continue

        for row in sheet.iter_rows(min_row=2):
            content_title = get_cell(sheet, row, "content title")
            if content_title.value:
                content_title.value = re.sub(r"\W+", "_", content_title.value.strip())


def add_english_keywords(workbook):
    """
    Goes through all of the non-english sheets, and adds the english keywords
    """
    english_keywords = get_keywords(workbook["English master"])
    for sheet in workbook:
        if sheet.title.strip().lower() in (
            "language codes",
            "importinfo",
            "english master",
            "sepedi (sa)",
        ):
            continue

        for row in sheet.iter_rows(min_row=2):
            keyword_cell = get_cell(sheet, row, "automation")
            keywords = (keyword_cell.value or "").split(",")
            language = get_cell(sheet, row, "language").value
            content_title = (get_cell(sheet, row, "content title").value or "").strip()
            # Skip empty row
            if not content_title:
                continue
            if content_title.startswith(language):
                content_title = content_title[len(language) + 1 :]
            if content_title not in english_keywords:
                error(f"Missing english content {content_title}, sheet: {sheet.title}")
                # get content title by row number if missing
                content_title = workbook["English master"][f"A{row[0].row}"].value[4:]
            for keyword in english_keywords[content_title]:
                if keyword not in keywords:
                    keywords.append(keyword)
            keywords = [k for k in keywords if k]
            keyword_cell.value = ",".join(keywords)


def check_content_length(workbook):
    for sheet in workbook:
        if sheet.title.strip().lower() in ("language codes", "importinfo"):
            continue

        for row in sheet.iter_rows(min_row=2):
            content = get_cell(sheet, row, "content").value or ""
            if len(content) > 4096:
                error(
                    f"Content too long: {get_cell(sheet, row, 'content title').value}; "
                    f"sheet: {sheet.title}"
                )


def add_missing_content(workbook):
    """
    If content for a specific language is missing, fill in using the English content
    """
    english_content = get_content(workbook["English master"])
    for sheet in workbook:
        if sheet.title.strip().lower() in (
            "language codes",
            "importinfo",
            "english master",
            "sepedi (sa)",
        ):
            continue

        for row in sheet.iter_rows(min_row=2):
            content_cell = get_cell(sheet, row, "content")
            content = (content_cell.value or "").strip()
            content_title = (get_cell(sheet, row, "content title").value or "").strip()
            language = get_cell(sheet, row, "language").value
            # Skip empty rows
            if not content_title:
                continue
            if content_title.startswith(language):
                content_title = content_title[len(language) + 1 :]
            if not content:
                content_cell.value = english_content[content_title]


if __name__ == "__main__":
    workbook = load_workbook(FILENAME)
    clean_language(workbook)
    clean_content_title(workbook)
    clean_keywords(workbook)
    add_english_keywords(workbook)
    check_content_length(workbook)
    add_missing_content(workbook)
    if is_error:
        raise Exception("There were errors, not saving")
    workbook.save(f"2{FILENAME}")
