import io
from contextlib import redirect_stdout
from unittest import TestCase

from openpyxl import Workbook

from process_sheet import (
    add_english_keywords,
    add_missing_content,
    base_emoji,
    check_content_length,
    clean_content_title,
    clean_keywords,
    clean_language,
    get_cell,
    get_index,
    get_keywords,
)


class TestProcessSheet(TestCase):
    def test_base_emoji(self):
        """
        Should return an emoji without any modifiers, if emoji
        """
        self.assertEqual(base_emoji("not emoji\U0001f3fb"), "not emoji\U0001f3fb")
        self.assertEqual(base_emoji("👍"), "👍")
        self.assertEqual(base_emoji("👍🏿"), "👍")

    def test_get_index(self):
        """
        Returns the index of the first matching header name
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "First Column"
        ws["B1"] = "Second Column "
        # Test empty column handling
        ws["D1"] = "Third Column"

        self.assertEqual(get_index(ws, "first column"), 0)
        self.assertEqual(get_index(ws, "second column"), 1)
        self.assertEqual(
            get_index(ws, "third_column", "third column", "first column"), 3
        )
        with self.assertRaises(AttributeError):
            get_index(ws, "no", "matching", "columns")

    def test_get_cell(self):
        """
        Returns the cell in the row according to the name of the cell
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Automation Title "
        ws["A2"] = "eng_welcome"
        ws["B1"] = "content"
        ws["B2"] = "Welcome!"
        ws["C1"] = "Language"
        ws["C2"] = "eng"
        ws["E1"] = "Automation"
        ws["E2"] = "hi,hello"

        self.assertEqual(get_cell(ws, ws[2], "content title"), ws["A2"])
        self.assertEqual(get_cell(ws, ws[2], "content"), ws["B2"])
        self.assertEqual(get_cell(ws, ws[2], "language"), ws["C2"])
        self.assertEqual(get_cell(ws, ws[2], "automation"), ws["E2"])

    def test_get_keywords(self):
        """
        Returns a map of the content title without language, to the list of keywords
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Automation Title "
        ws["A2"] = "eng_welcome"
        ws["A3"] = "eng_menu"
        ws["A4"] = "mother"
        ws["C1"] = "Language"
        ws["C2"] = "eng"
        ws["C3"] = "eng"
        ws["C4"] = "eng"
        ws["E1"] = "Automation"
        ws["E2"] = "hi,hello"
        # No keywords for third row, test no keyword handling
        ws["E4"] = "mother,motherhood"

        self.assertEqual(
            get_keywords(ws),
            {
                "welcome": ["hi", "hello"],
                "menu": [""],
                "mother": ["mother", "motherhood"],
            },
        )

    def test_clean_keywords(self):
        """
        Should clean keywords by:
        - Removing whitespace
        - Removing empty and duplicate
        - Removing emoji modifiers
        - Converting None to empty string
        - Converting float to trucated string
        - Converting int to string
        Shouldn't log an error for:
        - Duplicate keywords for myths
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Automation"
        ws["B1"] = "content title"
        ws["A2"] = "test1, whitespace"
        ws["B2"] = "test_whitespace"
        ws["A3"] = "test2,test2,duplicate"
        ws["B3"] = "test_duplicate"
        ws["A4"] = "test3,,empty"
        ws["B4"] = "test_empty"
        ws["A5"] = "test4,👍🏿,emoji"
        ws["B5"] = "test_emoji"
        ws["A6"] = None
        ws["B6"] = "test_none"
        ws["A7"] = 6.2
        ws["B7"] = "test_float"
        ws["A8"] = 5
        ws["B8"] = "test_int"
        ws["A9"] = "myths"
        ws["B9"] = "test_myths1"
        ws["A10"] = "myths"
        ws["B10"] = "test_myths2"

        output = io.StringIO()
        with redirect_stdout(output):
            clean_keywords(wb)

        self.assertEqual(ws["A2"].value, "test1,whitespace")
        self.assertEqual(ws["A3"].value, "test2,duplicate")
        self.assertEqual(ws["A4"].value, "test3,empty")
        self.assertEqual(ws["A5"].value, "test4,👍,emoji")
        self.assertEqual(ws["A6"].value, "")
        self.assertEqual(ws["A7"].value, "6")
        self.assertEqual(ws["A8"].value, "5")
        self.assertEqual(ws["A9"].value, "myths")
        self.assertEqual(ws["A10"].value, "myths")

        self.assertEqual(output.getvalue(), "")

    def test_clean_keywords_errors(self):
        """
        Should log an error when
        - Keyword that contains more than just an emoji
        - Keyword duplicate across rows in same sheet
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Automation"
        ws["B1"] = "content title"
        ws["A2"] = "test,duplicate"
        ws["B2"] = "test_duplicate1"
        ws["A3"] = "test,duplicate"
        ws["B3"] = "test_duplicate2"
        ws["A4"] = "test3,👍foo,emoji"
        ws["B4"] = "test_emoji"

        output = io.StringIO()
        with redirect_stdout(output):
            clean_keywords(wb)

        self.assertEqual(ws["A2"].value, "test,duplicate")
        self.assertEqual(ws["A3"].value, "test,duplicate")
        self.assertEqual(ws["A4"].value, "test3,👍foo,emoji")

        self.assertEqual(
            output.getvalue(),
            "\n".join(
                [
                    "Duplicate keyword test; sheet: Sheet",
                    "Duplicate keyword duplicate; sheet: Sheet",
                    "Invalid keyword, more than just emoji: 👍foo; sheet: Sheet",
                    "",
                ]
            ),
        )

    def test_content_length(self):
        """
        Should log an error when the content length is too long
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Content"
        ws["A2"] = "a" * 4097
        ws["B1"] = "Content Title"
        ws["B2"] = "test"

        output = io.StringIO()
        with redirect_stdout(output):
            check_content_length(wb)

        self.assertEqual(
            output.getvalue(),
            "\n".join(
                [
                    "Content too long: test; sheet: Sheet",
                    "",
                ]
            ),
        )

    def test_add_english_keywords(self):
        """
        Should add english keywords to existing keywords
        """
        wb = Workbook()
        eng_ws = wb.active
        eng_ws.title = "English master"
        eng_ws["A1"] = "Content title"
        # Row 2 is empty, to test empty row handling
        eng_ws["A3"] = "eng_test"
        eng_ws["B1"] = "Automation"
        eng_ws["B3"] = "test1,test2"
        eng_ws["C1"] = "Language"
        eng_ws["C3"] = "eng"

        ws = wb.create_sheet(title="Portugese")
        ws["A1"] = "Content title"
        # Row 2 is empty, to test empty row handling
        ws["A3"] = "por_test"
        ws["B1"] = "Automation"
        ws["B3"] = "test2,test3"
        ws["C1"] = "Language"
        ws["C3"] = "por"

        output = io.StringIO()
        with redirect_stdout(output):
            add_english_keywords(wb)

        self.assertEqual(ws["B3"].value, "test2,test3,test1")
        self.assertEqual(output.getvalue(), "")

    def test_add_english_keywords_content_title(self):
        """
        Should log the error and use row numbers to find english keyword
        """
        wb = Workbook()
        eng_ws = wb.active
        eng_ws.title = "English master"
        eng_ws["A1"] = "Content title"
        eng_ws["A2"] = "eng_test1"
        eng_ws["B1"] = "Automation"
        eng_ws["B2"] = "test1"
        eng_ws["C1"] = "Language"
        eng_ws["C2"] = "eng"

        ws = wb.create_sheet(title="Portuguese")
        ws["A1"] = "Content title"
        ws["A2"] = "por_not_test1"
        ws["B1"] = "Automation"
        ws["B2"] = "test2"
        ws["C1"] = "Language"
        ws["C2"] = "por"

        output = io.StringIO()
        with redirect_stdout(output):
            add_english_keywords(wb)

        self.assertEqual(ws["B2"].value, "test2,test1")
        self.assertEqual(
            output.getvalue(), "Missing english content not_test1, sheet: Portuguese\n"
        )

    def test_clean_language(self):
        """
        Should fill in any blank language fields using the previous row's language value
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Content title"
        ws["A2"] = "eng_test1"
        ws["A3"] = "eng_test2"
        ws["C1"] = "Language"
        ws["C2"] = "eng"

        clean_language(wb)

        self.assertEqual(ws["C3"].value, "eng")

    def test_clean_content_title(self):
        """
        Should do the following to content titles:
        - Remove leading and trailing whitespace
        - Replace non-word characters with _
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Content title"
        ws["A2"] = " leading and trailing whitespace "
        ws["A3"] = "non-word - characters"

        clean_content_title(wb)

        self.assertEqual(ws["A2"].value, "leading_and_trailing_whitespace")
        self.assertEqual(ws["A3"].value, "non_word_characters")

    def test_add_missing_content(self):
        """
        Should add english content where content is missing in other languages
        """
        wb = Workbook()
        eng_ws = wb.active
        eng_ws.title = "English master"
        eng_ws["A1"] = "Content title"
        # Row 2 is empty, to test empty row handling
        eng_ws["A3"] = "eng_test"
        eng_ws["B1"] = "Content"
        eng_ws["B3"] = "English Content"
        eng_ws["C1"] = "Language"
        eng_ws["C3"] = "eng"

        ws = wb.create_sheet(title="Portugese")
        ws["A1"] = "Content title"
        # Row 2 is empty, to test empty row handling
        ws["A3"] = "por_test"
        ws["A4"] = "existing"
        ws["B1"] = "Content"
        ws["B4"] = "portuguese content"
        ws["C1"] = "Language"
        ws["C3"] = "por"
        ws["C4"] = "por"

        output = io.StringIO()
        with redirect_stdout(output):
            add_missing_content(wb)

        self.assertEqual(ws["B3"].value, "English Content")
        self.assertEqual(ws["B4"].value, "portuguese content")
        self.assertEqual(output.getvalue(), "")
